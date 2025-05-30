import datetime
import io

from aiogram import types
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.db.repositories import MessageRepository
from src.keyboards.inline import callbacks, user


async def report_command(msg: types.Message) -> None:
    await msg.answer(
        "Выберите временной промежуток для отчета:",
        reply_markup=user.report.ReportButtons().main(),
    )


async def generate_report(
    cb: types.CallbackQuery,
    callback_data: callbacks.ReportCallbackFactory,
    message_repository: MessageRepository,
) -> None:
    # Определяем временной промежуток
    now = datetime.datetime.now(datetime.timezone.utc)

    period = callback_data.period
    if period == "day":
        start_date = now - datetime.timedelta(days=1)
    elif period == "week":
        start_date = now - datetime.timedelta(weeks=1)
    elif period == "month":
        start_date = now - datetime.timedelta(days=30)
    else:
        await cb.message.answer("Пожалуйста, выберите корректный временной промежуток.")
        return

    start_date = start_date.replace(tzinfo=None)

    # Получаем данные из базы данных
    messages = await message_repository.get_with_details(cb.from_user.id, start_date)

    # Создаем Excel-файл
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчет"

    # Определяем стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092",
        end_color="366092",
        fill_type="solid",
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    content_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Заголовки
    headers = [
        "ID сообщения",
        "Telegram ID сообщения",
        "Тип отправителя",
        "Содержание сообщения",
        "Количество токенов промпта",
        "Количество токенов ответа",
        "ID родительского сообщения",
        "ID треда",
        "ID инициатора треда",
        "Имя пользователя инициатора",
        "ID чата",
        "Название чата",
        "Активен ли чат",
        "ID темы",
        "Название темы",
        "Описание темы",
        "Промпт темы",
        "Оценка уверенности",
        "Дата и время создания",
        "Дата и время последней активности",
    ]

    # Устанавливаем заголовки и их стили
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Добавляем данные
    for row_num, message in enumerate(messages, 2):
        row_data = [
            message.id,
            message.telegram_message_id,
            message.sender_type,
            message.content,
            message.prompt_tokens,
            message.completion_tokens,
            message.parent_message_id,
            message.thread_id,
            message.thread.initiator_id,
            message.thread.initiator_username,
            message.thread.chat.telegram_chat_id,
            message.thread.chat.name,
            message.thread.chat.is_active,
            message.thread.topic.id,
            message.thread.topic.name,
            message.thread.topic.description,
            message.thread.topic.prompt,
            message.thread.confidence_score,
            message.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            message.thread.last_activity_at.strftime("%Y-%m-%d %H:%M:%S"),
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = content_alignment
            cell.border = thin_border

    # Устанавливаем ширину столбцов
    column_widths = {
        "A": 10,  # ID сообщения
        "B": 15,  # Telegram ID сообщения
        "C": 15,  # Тип отправителя
        "D": 50,  # Содержание сообщения
        "E": 20,  # Количество токенов промпта
        "F": 20,  # Количество токенов ответа
        "G": 20,  # ID родительского сообщения
        "H": 10,  # ID треда
        "I": 15,  # ID инициатора треда
        "J": 20,  # Имя пользователя инициатора
        "K": 20,  # ID чата
        "L": 20,  # Название чата
        "M": 15,  # Активен ли чат
        "N": 10,  # ID темы
        "O": 20,  # Название темы
        "P": 60,  # Описание темы
        "Q": 80,  # Промпт темы
        "R": 15,  # Оценка уверенности
        "S": 20,  # Дата и время создания
        "T": 20,  # Дата и время последней активности
    }

    for col_letter, width in column_widths.items():
        sheet.column_dimensions[col_letter].width = width

    # Устанавливаем высоту строк
    sheet.row_dimensions[1].height = 30  # Заголовок
    for row in range(2, len(messages) + 2):
        sheet.row_dimensions[row].height = 60  # Максимальная высота для строк с данными

    # Фиксируем заголовок
    sheet.freeze_panes = "A2"

    # Сохраняем файл в память
    file_stream = io.BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    # Создаем BufferedInputFile
    input_file = types.BufferedInputFile(
        file=file_stream.getvalue(),
        filename="report.xlsx",  # Получаем содержимое файла
    )

    # Отправляем файл пользователю
    await cb.message.answer_document(
        document=input_file,
        caption="Ваш отчет готов!",
    )
